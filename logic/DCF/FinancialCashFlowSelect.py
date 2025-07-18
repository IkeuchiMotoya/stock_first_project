import os
import time
import re
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from bs4 import BeautifulSoup
from openpyxl import load_workbook

def get_issued_shares_and_close_yahoo(ticker):
    url = f"https://finance.yahoo.co.jp/quote/{ticker}"
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(url)
    time.sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()
    issued = ''
    close = ''
    # 発行済株式数（参考指標の枠を最優先、その次dlで探す）
    for span in soup.find_all("span", class_=re.compile("DataListItem__name__3RQJ")):
        if "発行済株式数" in span.text:
            # 直近のddで数字を取得
            parent = span.find_parent("dt")
            dd = parent.find_next_sibling("dd")
            m = re.search(r"([0-9,]+)", dd.text if dd else '')
            if m:
                issued = m.group(1).replace(",", "")
            break
    if not issued:
        # フォールバック（dlタグ）
        for dl in soup.find_all("dl"):
            dt = dl.find("dt")
            dd = dl.find("dd")
            if dt and "発行済株式数" in dt.text:
                m = re.search(r"([0-9,]+)", dd.text if dd else '')
                if m:
                    issued = m.group(1).replace(",", "")
                break
    # 前日終値
    for span in soup.find_all("span", class_=re.compile("DataListItem__name__3RQJ")):
        if "前日終値" in span.text:
            parent = span.find_parent("dt")
            dd = parent.find_next_sibling("dd")
            m = re.search(r"([0-9,]+)", dd.text if dd else '')
            if m:
                close = m.group(1).replace(",", "")
            break
    return issued, close

def get_yahoo_table(url):
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    driver.get(url)
    time.sleep(2)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()
    table = soup.find("table")
    if not table:
        return pd.DataFrame()
    rows = table.find_all("tr")
    data = []
    for row in rows:
        data.append([cell.text.strip() for cell in row.find_all(["td", "th"])])
    df = pd.DataFrame(data)
    df.columns = df.iloc[0]
    df = df.drop(df.index[0])
    return df.reset_index(drop=True)

def save_df_to_sheet(writer, df, sheet_name):
    df.to_excel(writer, sheet_name=sheet_name, index=False)

def set_column_as_text(filepath, sheet_name, col_letters):
    wb = load_workbook(filepath)
    ws = wb[sheet_name]
    for col in col_letters:
        for cell in ws[col]:
            if cell.row == 1:
                continue  # ヘッダーはスキップ
            cell.number_format = '@'
    wb.save(filepath)

def main():
    # === 銘柄コード取得 ===
    csv_path = r"data\input\通期業績の推移、指標の取得\競合判定結果.csv"
    df_csv = pd.read_csv(csv_path, encoding="utf-8")
    ticker_raw = str(df_csv.iloc[0, 0]).strip()
    ticker = ticker_raw if ticker_raw.endswith('.T') else f"{ticker_raw}.T"

    results_dir = "data/output/競合他社との通期業績比較"
    os.makedirs(results_dir, exist_ok=True)
    outpath = os.path.join(results_dir, f"ヤフーファイナンス_財務_キャッシュフロー.xlsx")

    # 1. 発行済株式数・前日終値
    issued, close = get_issued_shares_and_close_yahoo(ticker)
    df_issued = pd.DataFrame({
        "銘柄コード": [ticker],
        "発行済株式数": [issued],
        "前日終値": [close]
    })

    # 2. 財務
    url_fin = f"https://finance.yahoo.co.jp/quote/{ticker}/performance?styl=financials"
    df_fin = get_yahoo_table(url_fin)

    # 3. キャッシュフロー
    url_cf = f"https://finance.yahoo.co.jp/quote/{ticker}/performance?styl=cf"
    df_cf = get_yahoo_table(url_cf)

    # Excel出力
    with pd.ExcelWriter(outpath, engine="openpyxl") as writer:
        save_df_to_sheet(writer, df_issued, "発行済株式数")
        save_df_to_sheet(writer, df_fin, "財務")
        save_df_to_sheet(writer, df_cf, "キャッシュフロー")

    # 発行済株式数シートのB・C列を文字列に
    set_column_as_text(outpath, "発行済株式数", ["B", "C"])

    print(f"出力完了: {outpath}")

if __name__ == "__main__":
    main()

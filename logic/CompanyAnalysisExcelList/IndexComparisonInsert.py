import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

# === パス定義 ===
csv_path = r"data\input\通期業績の推移、指標の取得\競合判定結果.csv"
indicator_path = r"data/output/競合他社との通期業績比較/指標比較.xlsx"
output_root = r"data/output/分析"

print(f"判定CSV読込: {csv_path}")
df_csv = pd.read_csv(csv_path, encoding="utf-8")

print(f"指標比較Excel読込: {indicator_path}")
df_indicator = pd.read_excel(indicator_path, header=0)
headers = df_indicator.columns.tolist()
data_rows = df_indicator.values.tolist()  # すべてのデータ行（2次元リスト）

# 2行目（インデックス1）の1・2列目で対象決定（ここは1行目でOKならインデックス0で）
row = df_csv.iloc[0]
folder_key = f"{row[0]}_{row[1]}"
target_xlsx = f"{folder_key}_企業分析.xlsx"
folder_path = os.path.join(output_root, folder_key)
target_xlsx_path = os.path.join(folder_path, target_xlsx)

print(f"\n対象フォルダ: {folder_path}")
print(f"対象Excel: {target_xlsx_path}")

# Excel書き込み処理
if os.path.exists(target_xlsx_path):
    print("→ Excelファイル発見、処理開始...")
    wb = openpyxl.load_workbook(target_xlsx_path)
    if "他社比較" in wb.sheetnames:
        ws = wb["他社比較"]
        print("→ '他社比較'シート有。")
    else:
        ws = wb.create_sheet("他社比較")
        print("→ '他社比較'シート新規作成。")

    start_col = 2  # B列
    start_row = 4  # 4行目

    # 1行目（ヘッダー行）埋め込み
    print(f"→ B4から横にヘッダー埋め込み: {headers}")
    for idx, header in enumerate(headers):
        col_letter = get_column_letter(start_col + idx)
        ws[f"{col_letter}{start_row}"] = header

    # 2行目以降にデータ行を順次書き込み
    for row_idx, data_row in enumerate(data_rows):
        excel_row = start_row + 1 + row_idx  # ヘッダーの下から
        for col_idx, cell_value in enumerate(data_row):
            col_letter = get_column_letter(start_col + col_idx)
            ws[f"{col_letter}{excel_row}"] = cell_value

    wb.save(target_xlsx_path)
    print("→ 書き込み・保存完了。")
else:
    print(f"× 指定Excelファイルが存在しません: {target_xlsx_path}")

print("\n処理終了")
